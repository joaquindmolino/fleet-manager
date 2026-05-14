"""Exportación de viajes a Excel (XLSX) usando openpyxl."""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "Nombre del viaje",
    "Documento asociado",
    "Cliente",
    "Conductor",
    "Vehículo",
    "Estado",
    "Fecha de creación",
    "Fecha programada",
    "Inicio real",
    "Fin real",
    "Origen",
    "Destino",
    "Paradas",
    "Km inicial",
    "Km final",
    "Km recorridos",
    "Observaciones",
]


def build_trips_xlsx(rows: list[dict]) -> bytes:
    """Construye un XLSX con las filas dadas. Cada fila es un dict cuyas keys
    deben coincidir con HEADERS."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Viajes"

    # Header row con estilo
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    # Datos
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(HEADERS, start=1):
            value = row.get(h)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if isinstance(value, datetime):
                cell.number_format = "dd/mm/yyyy hh:mm"

    # Anchos razonables por columna
    widths = {
        "Nombre del viaje": 24, "Documento asociado": 18, "Cliente": 22,
        "Conductor": 22, "Vehículo": 12, "Estado": 12,
        "Fecha de creación": 18, "Fecha programada": 18,
        "Inicio real": 18, "Fin real": 18,
        "Origen": 30, "Destino": 30,
        "Paradas": 9, "Km inicial": 12, "Km final": 12, "Km recorridos": 14,
        "Observaciones": 40,
    }
    for col_idx, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(h, 16)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
